# -*- coding: utf-8 -*-
"""
build_mapping_review.py
สร้างไฟล์ Excel ให้ผู้ใช้รีวิวการ mapping HC group โดยแยกตามระดับความมั่นใจ
Output: Mapping_Review.xlsx
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(r"C:\Users\nakit\OneDrive\Desktop\Healthier Choice")
INPUT = BASE / "evaluation_result.xlsx"
OUT = BASE / "Mapping_Review.xlsx"

HC_GROUPS = [
    "meal", "beverage", "seasoning", "dairy", "instant_food",
    "snack", "ice_cream", "fat_oil", "bread", "cereal",
    "bakery", "small_meal", "fish_seafood", "meat_product",
    "milk_alternative", "OUT_OF_SCOPE",
]

# กรอกค่าว่างให้กลายเป็น string ของ NaN
def fillna_str(df):
    return df.fillna("")

# ── โหลด ──
print("Loading evaluation_result.xlsx ...")
df = pd.read_excel(INPUT, sheet_name="Result")
print(f"  Total: {len(df)} products")

# คอลัมน์ที่ใช้
keep_cols = [
    "ลำดับ", "ชื่อลูกค้า", "จังหวัด", "ผลิตภัณฑ์", "ประเภท (อย.)",
    "HC_Group", "HC_Subgroup", "Mapping_Confidence", "Mapping_Keyword",
    "Criteria", "Failed_Nutrients", "Missing_Nutrients", "Note",
]
view = df[keep_cols].copy()

# เพิ่มช่องให้รีวิว
view.insert(6, "HC_Group_แก้ไข", "")
view.insert(7, "หมายเหตุรีวิว", "")

# ── ระดับความมั่นใจ ──
LEVELS = {
    "🚨_เร่งด่วน_Conf_1": (1, 1, "ต้องรีวิว — match จาก keyword กว้างมาก เช่น 'เครื่องดื่ม' / 'ขนม' / 'อาหาร'"),
    "⚠️_รีวิว_Conf_2":   (2, 2, "ควรรีวิว — match จาก keyword ทั่วไป อาจไม่แม่น"),
    "🔎_ตรวจ_Conf_3":    (3, 3, "ตรวจสุ่ม — match จาก keyword ค่อนข้างทั่วไป"),
    "✅_Conf_4":          (4, 4, "ค่อนข้างมั่นใจ — สุ่มดูเพื่อตรวจสอบ"),
    "✓_Conf_5":           (5, 5, "มั่นใจสูง — match keyword เฉพาะเจาะจง"),
    "❓_Unmapped":        (0, 0, "Map ไม่ได้ — ต้องระบุกลุ่ม HC เอง"),
}

# ── Workbook ──
print(f"Building {OUT.name} ...")
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "0_วิธีรีวิว"

# Sheet 0: คำอธิบายและสรุป
ws_summary["A1"] = "📋 คู่มือรีวิว Mapping ประเภทอาหาร 1,358 ผลิตภัณฑ์ → กลุ่ม HC"
ws_summary["A1"].font = Font(name="Tahoma", size=14, bold=True, color="1F4E78")
ws_summary.merge_cells("A1:F1")

ws_summary["A3"] = "วิธีรีวิว"
ws_summary["A3"].font = Font(name="Tahoma", size=12, bold=True)
ws_summary.merge_cells("A3:F3")

instructions = [
    "1) เริ่มรีวิวจาก sheet '🚨_เร่งด่วน_Conf_1' (จำเป็นต้องตรวจ — match กว้างมาก)",
    "2) คอลัมน์ 'HC_Group_แก้ไข' = ใส่กลุ่มที่ถูกต้อง (มี dropdown ให้เลือก)",
    "    • ถ้าปล่อยว่าง = ใช้ค่าที่ระบบจัดให้ในคอลัมน์ HC_Group",
    "    • ถ้าใส่ค่า = จะ override การ mapping สำหรับผลิตภัณฑ์นั้น",
    "    • ใส่ 'OUT_OF_SCOPE' หากเป็นผลิตภัณฑ์นอกขอบเขต HC (เช่น น้ำผึ้ง น้ำดื่ม)",
    "3) คอลัมน์ 'หมายเหตุรีวิว' = เขียนเหตุผลหรือบันทึกอย่างย่อ",
    "4) เมื่อรีวิวเสร็จ บอกผม จะนำการแก้ไขใส่กลับเข้าระบบและรันใหม่",
    "",
    "💡 ระดับ Confidence (0-5)",
    "   5 = match keyword เฉพาะเจาะจง เช่น 'ไอศกรีม' 'มายองเนส'",
    "   4 = match keyword ค่อนข้างเฉพาะ",
    "   3 = match keyword ทั่วไป",
    "   2 = match keyword กว้าง",
    "   1 = match keyword กว้างมาก เช่น 'เครื่องดื่ม' 'ขนม' 'อาหาร'",
    "   0 = ไม่สามารถ map ได้",
]
for i, line in enumerate(instructions, start=4):
    ws_summary.cell(row=i, column=1, value=line)
    ws_summary.cell(row=i, column=1).font = Font(name="Tahoma", size=11)
    ws_summary.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ตารางสรุปจำนวน
ws_summary["A22"] = "📊 สรุปจำนวนผลิตภัณฑ์แต่ละระดับ"
ws_summary["A22"].font = Font(name="Tahoma", size=12, bold=True)
ws_summary["A24"] = "Sheet"
ws_summary["B24"] = "Confidence"
ws_summary["C24"] = "จำนวน"
ws_summary["D24"] = "ลำดับความสำคัญ"
ws_summary["E24"] = "คำอธิบาย"
for col in range(1, 6):
    c = ws_summary.cell(row=24, column=col)
    c.font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

row_i = 25

# ── สร้าง sheet แต่ละระดับ ──
print()
print(f"{'Sheet':<35}  Confidence  จำนวน")
print("-" * 60)
for sheet_name, (min_c, max_c, desc) in LEVELS.items():
    subset = view[(view["Mapping_Confidence"] >= min_c) &
                  (view["Mapping_Confidence"] <= max_c)].copy()
    n = len(subset)
    print(f"{sheet_name:<35}  {min_c}-{max_c}        {n}")
    if n == 0:
        continue

    # สรุปลงใน sheet 0
    priority_emoji = {1: "🚨 ด่วน", 2: "⚠️ สูง", 3: "🔎 กลาง", 4: "✅ ต่ำ", 5: "✓ ไม่ต้อง"}.get(min_c, "❓")
    ws_summary.cell(row=row_i, column=1, value=sheet_name)
    ws_summary.cell(row=row_i, column=2, value=f"{min_c}" if min_c == max_c else f"{min_c}-{max_c}")
    ws_summary.cell(row=row_i, column=3, value=n)
    ws_summary.cell(row=row_i, column=4, value=priority_emoji)
    ws_summary.cell(row=row_i, column=5, value=desc)
    for col in range(1, 6):
        c = ws_summary.cell(row=row_i, column=col)
        c.font = Font(name="Tahoma", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    row_i += 1

    # สร้าง sheet
    ws = wb.create_sheet(sheet_name)
    ws.append(list(subset.columns))
    # Header style
    for col in range(1, len(subset.columns) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Data rows
    for _, r in subset.iterrows():
        ws.append(list(r.values))

    # Style เนื้อหา
    for r_idx in range(2, len(subset) + 2):
        for c_idx in range(1, len(subset.columns) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Tahoma", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # Column widths
    widths = {
        "ลำดับ": 7, "ชื่อลูกค้า": 22, "จังหวัด": 12, "ผลิตภัณฑ์": 45,
        "ประเภท (อย.)": 22, "HC_Group": 14, "HC_Group_แก้ไข": 16,
        "หมายเหตุรีวิว": 28, "HC_Subgroup": 18, "Mapping_Confidence": 8,
        "Mapping_Keyword": 18, "Criteria": 9, "Failed_Nutrients": 25,
        "Missing_Nutrients": 22, "Note": 25,
    }
    for i, col_name in enumerate(subset.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 12)

    # Freeze + Filter
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    # Data Validation dropdown ใน "HC_Group_แก้ไข" (column G = index 7)
    col_letter = get_column_letter(7)  # HC_Group_แก้ไข
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(HC_GROUPS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.add(f"{col_letter}2:{col_letter}{len(subset)+1}")
    ws.add_data_validation(dv)

    # Highlight Confidence column (สีอ่อนตามระดับ)
    conf_col = get_column_letter(list(subset.columns).index("Mapping_Confidence") + 1)
    if min_c == 1:
        fill = PatternFill("solid", fgColor="FFCDD2")  # แดงอ่อน
    elif min_c == 2:
        fill = PatternFill("solid", fgColor="FFE0B2")  # ส้มอ่อน
    elif min_c == 3:
        fill = PatternFill("solid", fgColor="FFF9C4")  # เหลืองอ่อน
    elif min_c == 4:
        fill = PatternFill("solid", fgColor="DCEDC8")  # เขียวอ่อน
    elif min_c == 5:
        fill = PatternFill("solid", fgColor="C8E6C9")  # เขียว
    else:
        fill = PatternFill("solid", fgColor="ECEFF1")  # เทา
    for r_idx in range(2, len(subset) + 2):
        ws.cell(row=r_idx, column=list(subset.columns).index("Mapping_Confidence") + 1).fill = fill

# ── Style sheet 0 ──
ws_summary.column_dimensions["A"].width = 35
ws_summary.column_dimensions["B"].width = 12
ws_summary.column_dimensions["C"].width = 10
ws_summary.column_dimensions["D"].width = 14
ws_summary.column_dimensions["E"].width = 60
ws_summary.column_dimensions["F"].width = 10
ws_summary.row_dimensions[1].height = 24
ws_summary.row_dimensions[22].height = 22

# ── Save ──
wb.save(OUT)
print()
print(f"✓ สร้างไฟล์ {OUT.name} สำเร็จ ({OUT.stat().st_size/1024:.1f} KB)")
print(f"  → {OUT}")
print()
print("🎯 ขั้นตอนรีวิว (เรียงตามลำดับความสำคัญ):")
print("   1. '🚨_เร่งด่วน_Conf_1' — เริ่มที่นี่ (match กว้างมาก)")
print("   2. '⚠️_รีวิว_Conf_2'")
print("   3. '🔎_ตรวจ_Conf_3'")
print("   4. '✅_Conf_4' (สุ่มดู)")
print("   5. '✓_Conf_5' (ไม่ต้องรีวิว เว้นแต่อยากตรวจสุ่ม)")
